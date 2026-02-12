# extract_transform_erp_table_sales_gp.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Sequence, Tuple, Dict, List

import numpy as np
import pandas as pd


@dataclass(frozen=True)
class SalesGPMergeConfig:
    input_filename: str = "[MK P&L] Выручка и себестоимость ГП и товары.xlsx"
    additional_filenames: Tuple[str, ...] = (
        "[MK P&L] Выручка и себестоимость 45",
        "[MK P&L] Выручка услуги",
    )

    base_dir: Optional[Path] = None
    sheet_name: Optional[str] = None

    doc_col: str = "Товары.Ссылка"
    doc_type_substring: str = "Реализация товаров и услуг"
    key_cols: Tuple[str, ...] = ("Товары.Ссылка", "Товары.Ссылка.Контрагент", "Товары.Номенклатура")
    volume_col: str = "Объем отгрузки, тн"
    cost_col: str = "Сумма СС"
    eps: float = 1e-12

    sku_group_col: str = "SKU GROUP NAME"
    fin_group_col_main: str = "Документ.Заказ клиента.Группа фин. учета расчетов"
    fin_group_col_services: str = "Товары.Заказ клиента.Группа фин. учета расчетов"

    merged_output_filename: str = "sales_gp_merged.xlsx"
    merged_output_sheet: str = "data"

    # ✅ куда сохранять: True -> рядом с main excel, False -> рядом со скриптом
    save_next_to_main_excel: bool = False


class NumericParser:
    @staticmethod
    def to_float(series: pd.Series) -> pd.Series:
        s = series.astype("string")
        s = s.str.replace("\u00A0", "", regex=False)
        s = s.str.replace(" ", "", regex=False)
        s = s.str.replace(",", ".", regex=False)
        return pd.to_numeric(s, errors="coerce").fillna(0.0)


class ExcelPathResolver:
    def __init__(self, cfg: SalesGPMergeConfig):
        self.cfg = cfg

    @staticmethod
    def _ensure_xlsx_name(name: str) -> List[str]:
        name = name.strip()
        if name.lower().endswith((".xlsx", ".xls")):
            return [name]
        return [f"{name}.xlsx", name]

    def resolve_main_input_path(self) -> Path:
        fname = self.cfg.input_filename
        candidates: List[Path] = []

        if self.cfg.base_dir is not None:
            candidates.append(self.cfg.base_dir / fname)

        candidates.append(Path(__file__).resolve().parent / fname)
        candidates.append(Path(r"Z:\01 Администрация\Кобычев\МК P&L") / fname)

        for p in candidates:
            if p.exists():
                return p

        tried = "\n".join(str(p) for p in candidates)
        raise FileNotFoundError(f"Не найден входной Excel. Пробовал пути:\n{tried}")

    def resolve_additional_paths(self, base_dir: Path) -> List[Path]:
        paths: List[Path] = []
        for raw in self.cfg.additional_filenames:
            found = None
            for nm in self._ensure_xlsx_name(raw):
                cand = base_dir / nm
                if cand.exists():
                    found = cand
                    break
            if found is None:
                tried = "\n".join(str(base_dir / nm) for nm in self._ensure_xlsx_name(raw))
                raise FileNotFoundError(f"Не найден дополнительный Excel: '{raw}'. Пробовал:\n{tried}")
            paths.append(found)
        return paths

    def output_path(self, main_excel_path: Path) -> Path:
        if self.cfg.save_next_to_main_excel:
            return main_excel_path.parent / self.cfg.merged_output_filename
        return Path(__file__).resolve().parent / self.cfg.merged_output_filename


class ExcelIO:
    @staticmethod
    def detect_first_sheet(path: Path) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return wb.sheetnames[0]
        finally:
            wb.close()

    @staticmethod
    def read_excel(path: Path, sheet_name: Optional[str] = None) -> tuple[pd.DataFrame, str]:
        use_sheet = sheet_name or ExcelIO.detect_first_sheet(path)
        df = pd.read_excel(path, sheet_name=use_sheet, engine="openpyxl")
        return df, use_sheet

    @staticmethod
    def write_excel(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            wb.save(path)
        except Exception:
            pass


class TableUnionByBaseColumns:
    @staticmethod
    def union(base_df: pd.DataFrame, other_dfs: Sequence[pd.DataFrame]) -> pd.DataFrame:
        base_cols = list(base_df.columns)
        aligned: List[pd.DataFrame] = [base_df.copy()]
        for odf in other_dfs:
            aligned.append(odf.reindex(columns=base_cols))
        return pd.concat(aligned, axis=0, ignore_index=True)


class ServicesPreprocessor:
    def __init__(self, *, fin_group_col_main: str, fin_group_col_services: str, sku_group_col: str):
        self.fin_group_col_main = fin_group_col_main
        self.fin_group_col_services = fin_group_col_services
        self.sku_group_col = sku_group_col
        self._map: Dict[str, str] = {
            "Покупатели (Продукция Фибратек)": "Фибратек",
            "Покупатели (Продукция ЛАТО)": "ЛАТО",
            "Покупатели (Продукция МК)": "Прочие",
            "Покупатели (Продукция СПБ)": "Прочие",
            "Покупатели (Продукция ТЕХПРОМ)": "Техпром",
            "Покупатели (Аренда)": "Прочие",
            "Покупатели (группа)": "Прочие",
        }

    @staticmethod
    def _norm_series(s: pd.Series) -> pd.Series:
        return (
            s.astype("string")
            .str.replace("\u00A0", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    def apply(self, df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()

        # normalize fin-group col name
        if self.fin_group_col_services in out.columns and self.fin_group_col_main not in out.columns:
            out = out.rename(columns={self.fin_group_col_services: self.fin_group_col_main})
        elif self.fin_group_col_services in out.columns and self.fin_group_col_main in out.columns:
            main = out[self.fin_group_col_main]
            srv = out[self.fin_group_col_services]
            out[self.fin_group_col_main] = main.where(main.notna(), srv)
            out = out.drop(columns=[self.fin_group_col_services])

        if self.sku_group_col not in out.columns:
            out[self.sku_group_col] = pd.NA

        if self.fin_group_col_main in out.columns:
            g = self._norm_series(out[self.fin_group_col_main])
            out[self.sku_group_col] = g.map(self._map).fillna("Прочие")
        else:
            out[self.sku_group_col] = "Прочие"

        return out


class RedistributeDuplicatedCostByVolume:
    def __init__(
        self,
        *,
        doc_col: str,
        doc_type_substring: str,
        key_cols: Sequence[str],
        volume_col: str,
        cost_col: str,
        eps: float,
    ):
        self.doc_col = doc_col
        self.doc_type_substring = doc_type_substring
        self.key_cols = list(key_cols)
        self.volume_col = volume_col
        self.cost_col = cost_col
        self.eps = float(eps)

        self._vol_tmp = "__vol_num"
        self._cost_tmp = "__cost_num"
        self._cnz_tmp = "__cost_nz_flag"
        self._cnzv_tmp = "__cost_nz_val"

    def run(self, df: pd.DataFrame) -> pd.DataFrame:
        need_cols = set(self.key_cols + [self.doc_col, self.volume_col, self.cost_col])
        missing = [c for c in need_cols if c not in df.columns]
        if missing:
            raise KeyError(f"В Excel не найдены обязательные колонки для фикса себестоимости: {missing}")

        out = df.copy()

        doc_mask = (
            out[self.doc_col]
            .astype("string")
            .str.contains(self.doc_type_substring, case=False, na=False)
        )

        vol = NumericParser.to_float(out[self.volume_col])
        cost = NumericParser.to_float(out[self.cost_col])

        out[self._vol_tmp] = vol
        out[self._cost_tmp] = cost

        nonzero = cost.abs() > self.eps
        out[self._cnz_tmp] = nonzero.astype("int8")
        out[self._cnzv_tmp] = cost.where(nonzero, np.nan)

        g = out.groupby(self.key_cols, dropna=False, sort=False)

        group_size = g[self._cost_tmp].transform("size").astype("int64")
        cost_nonzero_count = g[self._cnz_tmp].transform("sum").astype("int64")
        cost_nz_min = g[self._cnzv_tmp].transform("min")
        cost_nz_max = g[self._cnzv_tmp].transform("max")

        fix_mask = (
            doc_mask
            & (group_size > 1)
            & (cost_nonzero_count == group_size)
            & (cost_nz_min == cost_nz_max)
        )

        if not bool(fix_mask.any()):
            out[self.cost_col] = cost
            out.drop(columns=[self._vol_tmp, self._cost_tmp, self._cnz_tmp, self._cnzv_tmp], inplace=True, errors="ignore")
            return out

        total_cost = cost_nz_max.fillna(0.0)
        total_vol = g[self._vol_tmp].transform("sum")

        corrected = cost.to_numpy(copy=True)
        fix_np = fix_mask.to_numpy()

        total_vol_np = total_vol.to_numpy()
        vol_np = vol.to_numpy()
        total_cost_np = total_cost.to_numpy()
        group_size_np = group_size.to_numpy()

        vol_ok = np.abs(total_vol_np) > self.eps

        idx1 = fix_np & vol_ok
        corrected[idx1] = total_cost_np[idx1] * vol_np[idx1] / total_vol_np[idx1]

        idx2 = fix_np & (~vol_ok)
        corrected[idx2] = total_cost_np[idx2] / group_size_np[idx2]

        out[self.cost_col] = corrected
        out.drop(columns=[self._vol_tmp, self._cost_tmp, self._cnz_tmp, self._cnzv_tmp], inplace=True, errors="ignore")
        return out


def main():
    cfg = SalesGPMergeConfig()
    resolver = ExcelPathResolver(cfg)

    main_path = resolver.resolve_main_input_path()
    base_dir = main_path.parent
    add_paths = resolver.resolve_additional_paths(base_dir)
    out_path = resolver.output_path(main_path)

    df_main, _ = ExcelIO.read_excel(main_path, cfg.sheet_name)

    if cfg.sku_group_col not in df_main.columns:
        df_main[cfg.sku_group_col] = pd.NA

    svc_prep = ServicesPreprocessor(
        fin_group_col_main=cfg.fin_group_col_main,
        fin_group_col_services=cfg.fin_group_col_services,
        sku_group_col=cfg.sku_group_col,
    )

    other_dfs: List[pd.DataFrame] = []
    for p in add_paths:
        df_x, _ = ExcelIO.read_excel(p, cfg.sheet_name)

        # services mapping only for services file
        if "услуг" in p.stem.lower():
            df_x = svc_prep.apply(df_x)

        # normalize fin group column name
        if cfg.fin_group_col_services in df_x.columns and cfg.fin_group_col_main not in df_x.columns:
            df_x = df_x.rename(columns={cfg.fin_group_col_services: cfg.fin_group_col_main})

        if cfg.sku_group_col not in df_x.columns:
            df_x[cfg.sku_group_col] = pd.NA

        other_dfs.append(df_x)

    df_all = TableUnionByBaseColumns.union(df_main, other_dfs)

    df_all = RedistributeDuplicatedCostByVolume(
        doc_col=cfg.doc_col,
        doc_type_substring=cfg.doc_type_substring,
        key_cols=cfg.key_cols,
        volume_col=cfg.volume_col,
        cost_col=cfg.cost_col,
        eps=cfg.eps,
    ).run(df_all)

    ExcelIO.write_excel(out_path, df_all, cfg.merged_output_sheet)

    print("\n=== ГОТОВО: объединение 3 таблиц + фикс Сумма СС ===")
    print("Main   :", str(main_path))
    print("Output :", str(out_path))
    print("Rows   :", len(df_all))
    print("Cols   :", len(df_all.columns))


if __name__ == "__main__":
    main()
