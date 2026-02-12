# catboost_sales_gp_category_production.py
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, List, Dict, Any, Sequence, Tuple

import numpy as np
import pandas as pd

from sklearn.model_selection import GroupShuffleSplit
from sklearn.metrics import classification_report


@dataclass(frozen=True)
class CatBoostSkuConfig:
    labeled_filename: str = "sku_group_labeled.xlsx"
    merged_filename: str = "sales_gp_merged.xlsx"

    model_filename: str = "sku_group_model.cbm"
    meta_filename: str = "sku_group_model.meta.json"

    scored_filename: str = "sales_gp_scored.xlsx"
    scored_sheet: str = "data"

    active_learning_filename: str = "active_learning_to_label.xlsx"
    active_learning_sheet: str = "to_label"

    target_col: str = "SKU GROUP NAME"
    group_col: str = "Товары.Ссылка"
    fin_group_col_main: str = "Документ.Заказ клиента.Группа фин. учета расчетов"

    text_cols: Tuple[str, ...] = ("Товары.Номенклатура",)

    uncertainty_threshold: float = 0.60
    max_active_rows: int = 2500
    random_seed: int = 42
    drop_unique_ratio: float = 0.90


class ExcelIO:
    @staticmethod
    def detect_first_sheet(path: Path) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return wb.sheetnames[0]
        finally:
            wb.close()

    @staticmethod
    def read_excel(path: Path, sheet_name: Optional[str] = None) -> tuple[pd.DataFrame, str]:
        use_sheet = sheet_name or ExcelIO.detect_first_sheet(path)
        df = pd.read_excel(path, sheet_name=use_sheet, engine="openpyxl")
        return df, use_sheet

    @staticmethod
    def write_excel(path: Path, df: pd.DataFrame, sheet_name: str) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)


class FileFinder:
    @staticmethod
    def _safe_read_header(path: Path) -> Optional[List[str]]:
        try:
            df0 = pd.read_excel(path, nrows=1, engine="openpyxl")
            return list(df0.columns)
        except Exception:
            return None

    @staticmethod
    def find_labeled(base_dir: Path, cfg: CatBoostSkuConfig) -> Path:
        direct = base_dir / cfg.labeled_filename
        if direct.exists():
            return direct

        # try patterns
        candidates = sorted(base_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)

        # 1) name contains labeled
        for p in candidates:
            if "labeled" in p.stem.lower():
                cols = FileFinder._safe_read_header(p)
                if cols and cfg.target_col in cols:
                    return p

        # 2) any file that has target column
        for p in candidates:
            cols = FileFinder._safe_read_header(p)
            if cols and cfg.target_col in cols:
                return p

        raise FileNotFoundError(
            f"Не найден размеченный файл. Положи '{cfg.labeled_filename}' в папку проекта: {base_dir}"
        )

    @staticmethod
    def find_merged(base_dir: Path, cfg: CatBoostSkuConfig) -> Path:
        direct = base_dir / cfg.merged_filename
        if direct.exists():
            return direct

        candidates = sorted(base_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)

        # 1) name contains merged
        for p in candidates:
            if "merged" in p.stem.lower():
                cols = FileFinder._safe_read_header(p)
                if cols and (cfg.group_col in cols or cfg.target_col in cols):
                    return p

        # 2) file contains key columns typical for merged
        must_have_any = {"Товары.Ссылка", "Сумма СС", "Товары.Номенклатура"}
        for p in candidates:
            cols = FileFinder._safe_read_header(p)
            if cols and len(must_have_any.intersection(set(cols))) >= 2:
                return p

        raise FileNotFoundError(
            f"Не найден объединённый файл. Сначала запусти extract_transform_erp_table_sales_gp.py "
            f"и убедись что появился '{cfg.merged_filename}' в {base_dir}"
        )


class FeatureBuilder:
    def __init__(
        self,
        *,
        target_col: str,
        text_cols: Sequence[str],
        drop_unique_ratio: float,
    ):
        self.target_col = target_col
        self.text_cols = list(text_cols)
        self.drop_unique_ratio = float(drop_unique_ratio)

    @staticmethod
    def _norm_df(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        for c in out.columns:
            if out[c].dtype == "object":
                out[c] = out[c].astype("string")
        return out

    def pick_feature_cols(self, df: pd.DataFrame) -> List[str]:
        return [c for c in df.columns if c != self.target_col]

    def drop_id_like_cols(self, df: pd.DataFrame, feature_cols: List[str]) -> List[str]:
        n = len(df)
        keep: List[str] = []
        for c in feature_cols:
            if c in self.text_cols:
                keep.append(c)
                continue
            ser = df[c]
            if ser.dtype == "object" or str(ser.dtype).startswith("string"):
                uniq = ser.nunique(dropna=True)
                if n > 0 and (uniq / n) >= self.drop_unique_ratio:
                    continue
                keep.append(c)
            else:
                keep.append(c)
        return keep

    @staticmethod
    def ensure_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
        out = df.copy()
        for c in cols:
            if c not in out.columns:
                out[c] = pd.NA
        return out.reindex(columns=cols)

    def split_types(self, X: pd.DataFrame) -> tuple[List[str], List[str], List[str]]:
        text_cols = [c for c in self.text_cols if c in X.columns]
        cat_cols: List[str] = []
        num_cols: List[str] = []
        for c in X.columns:
            if c in text_cols:
                continue
            if pd.api.types.is_numeric_dtype(X[c]):
                num_cols.append(c)
            else:
                cat_cols.append(c)
        return text_cols, cat_cols, num_cols

    @staticmethod
    def fill_na(X: pd.DataFrame, text_cols: List[str], cat_cols: List[str], num_cols: List[str]) -> pd.DataFrame:
        out = X.copy()
        for c in cat_cols + text_cols:
            out[c] = out[c].astype("string").fillna("")
        for c in num_cols:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
        return out


class CatBoostSkuModel:
    def __init__(self, cfg: CatBoostSkuConfig):
        self.cfg = cfg

    def train(self, labeled_path: Path, model_path: Path, meta_path: Path) -> None:
        try:
            from catboost import CatBoostClassifier, Pool
        except Exception as e:
            raise RuntimeError("Нужен пакет catboost. Добавь 'catboost' в requirements.txt и установи.") from e

        df, _ = ExcelIO.read_excel(labeled_path)
        df = FeatureBuilder._norm_df(df)

        if self.cfg.target_col not in df.columns:
            raise KeyError(f"В размеченном файле нет '{self.cfg.target_col}'.")

        y = df[self.cfg.target_col].astype("string").fillna("Прочие")

        fb = FeatureBuilder(
            target_col=self.cfg.target_col,
            text_cols=self.cfg.text_cols,
            drop_unique_ratio=self.cfg.drop_unique_ratio,
        )
        feature_cols = fb.drop_id_like_cols(df, fb.pick_feature_cols(df))

        X = fb.ensure_cols(df, feature_cols)
        text_cols, cat_cols, num_cols = fb.split_types(X)
        X = fb.fill_na(X, text_cols, cat_cols, num_cols)

        # group split
        if self.cfg.group_col in df.columns:
            groups = df[self.cfg.group_col].astype("string").fillna("__NA__")
            splitter = GroupShuffleSplit(n_splits=1, test_size=0.2, random_state=self.cfg.random_seed)
            train_idx, valid_idx = next(splitter.split(X, y, groups=groups))
        else:
            rng = np.random.default_rng(self.cfg.random_seed)
            idx = np.arange(len(X))
            rng.shuffle(idx)
            cut = int(len(idx) * 0.8)
            train_idx, valid_idx = idx[:cut], idx[cut:]

        X_train, X_valid = X.iloc[train_idx], X.iloc[valid_idx]
        y_train, y_valid = y.iloc[train_idx], y.iloc[valid_idx]

        cat_features = [X.columns.get_loc(c) for c in cat_cols]
        text_features = [X.columns.get_loc(c) for c in text_cols]

        train_pool = Pool(X_train, label=y_train, cat_features=cat_features, text_features=text_features)
        valid_pool = Pool(X_valid, label=y_valid, cat_features=cat_features, text_features=text_features)

        model = CatBoostClassifier(
            loss_function="MultiClass",
            eval_metric="TotalF1",
            random_seed=self.cfg.random_seed,
            iterations=4000,
            learning_rate=0.06,
            depth=8,
            l2_leaf_reg=6.0,
            auto_class_weights="Balanced",
            od_type="Iter",
            od_wait=200,
            verbose=200,
        )
        model.fit(train_pool, eval_set=valid_pool, use_best_model=True)

        pred = model.predict(valid_pool).astype(str).ravel()
        print("\n=== Validation report ===")
        print(classification_report(y_valid.astype(str).to_numpy(), pred, digits=4))

        model.save_model(str(model_path))
        meta = {
            "feature_cols": feature_cols,
            "text_cols": text_cols,
            "cat_cols": cat_cols,
            "num_cols": num_cols,
            "target_col": self.cfg.target_col,
            "group_col": self.cfg.group_col,
            "model_type": "catboost",
        }
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

        print("\n=== Saved ===")
        print("Model:", model_path)
        print("Meta :", meta_path)

    @staticmethod
    def load(model_path: Path, meta_path: Path):
        from catboost import CatBoostClassifier
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        model = CatBoostClassifier()
        model.load_model(str(model_path))
        return model, meta

    @staticmethod
    def predict(model, meta: Dict[str, Any], df: pd.DataFrame) -> tuple[np.ndarray, np.ndarray]:
        from catboost import Pool

        feature_cols = meta["feature_cols"]
        text_cols = meta["text_cols"]
        cat_cols = meta["cat_cols"]
        num_cols = meta["num_cols"]

        X = df.copy()
        for c in feature_cols:
            if c not in X.columns:
                X[c] = pd.NA
        X = X.reindex(columns=feature_cols)

        for c in cat_cols + text_cols:
            X[c] = X[c].astype("string").fillna("")
        for c in num_cols:
            X[c] = pd.to_numeric(X[c], errors="coerce").fillna(0.0)

        cat_features = [X.columns.get_loc(c) for c in cat_cols]
        text_features = [X.columns.get_loc(c) for c in text_cols]
        pool = Pool(X, cat_features=cat_features, text_features=text_features)

        proba = model.predict_proba(pool)
        cls = np.array(model.classes_, dtype=object)
        pred_idx = np.argmax(proba, axis=1)
        pred = cls[pred_idx]
        conf = np.max(proba, axis=1)
        return pred.astype(str), conf.astype(float)


def build_lock_mask(df: pd.DataFrame, sku_col: str, fin_group_col: str) -> pd.Series:
    locked = pd.Series(False, index=df.index)
    if sku_col in df.columns and fin_group_col in df.columns:
        g = df[fin_group_col].astype("string").fillna("").str.strip()
        is_services = g.str.startswith("Покупатели", na=False)
        locked = df[sku_col].notna() & is_services
    return locked


def export_active_learning(
    df: pd.DataFrame,
    *,
    conf: np.ndarray,
    pred: np.ndarray,
    sku_col: str,
    threshold: float,
    max_rows: int,
    out_path: Path,
    sheet: str,
) -> None:
    cur = df[sku_col].astype("string") if sku_col in df.columns else pd.Series(pd.NA, index=df.index, dtype="string")
    cur = cur.where(cur.notna(), pd.NA)

    need = cur.isna().to_numpy() | (conf < threshold)

    sub = df.loc[need].copy()
    sub["__pred"] = pred[need]
    sub["__conf"] = conf[need]
    sub = sub.sort_values("__conf", ascending=True)

    if len(sub) > max_rows:
        sub = sub.head(max_rows)

    ExcelIO.write_excel(out_path, sub, sheet)
    print("\n=== Active learning file saved ===")
    print("Path:", out_path)
    print("Rows:", len(sub))


def main():
    cfg = CatBoostSkuConfig()

    parser = argparse.ArgumentParser()
    parser.add_argument("--train", action="store_true")
    parser.add_argument("--predict", action="store_true")
    parser.add_argument("--active-learning", action="store_true")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent

    labeled_path = FileFinder.find_labeled(base_dir, cfg)
    merged_path = FileFinder.find_merged(base_dir, cfg)

    model_path = base_dir / cfg.model_filename
    meta_path = base_dir / cfg.meta_filename
    scored_path = base_dir / cfg.scored_filename
    active_path = base_dir / cfg.active_learning_filename

    model_api = CatBoostSkuModel(cfg)

    if args.train:
        print("Labeled:", labeled_path)
        model_api.train(labeled_path, model_path, meta_path)
        return  # ✅ критично: после train не идём в predict

    if not args.predict and not args.active_learning:
        print("Ничего не выбрано. Используй --train или --predict (и опционально --active-learning).")
        return

    if not model_path.exists() or not meta_path.exists():
        raise FileNotFoundError("Модель не найдена. Сначала запусти: python catboost_sales_gp_category_production.py --train")

    df, _ = ExcelIO.read_excel(merged_path)
    if cfg.target_col not in df.columns:
        df[cfg.target_col] = pd.NA

    locked = build_lock_mask(df, cfg.target_col, cfg.fin_group_col_main)
    can_edit = ~locked

    model, meta = CatBoostSkuModel.load(model_path, meta_path)

    pred_all = np.array([""] * len(df), dtype=object)
    conf_all = np.zeros(len(df), dtype=float)

    idx = df.index[can_edit].to_numpy()
    if len(idx) > 0:
        pred, conf = CatBoostSkuModel.predict(model, meta, df.loc[idx])
        pred_all[idx] = pred
        conf_all[idx] = conf
        df.loc[idx, cfg.target_col] = pred

    ExcelIO.write_excel(scored_path, df, cfg.scored_sheet)
    print("\n=== Scored saved ===")
    print("Merged :", merged_path)
    print("Scored :", scored_path)
    print("Rows   :", len(df))

    if args.active_learning:
        export_active_learning(
            df,
            conf=conf_all,
            pred=pred_all,
            sku_col=cfg.target_col,
            threshold=cfg.uncertainty_threshold,
            max_rows=cfg.max_active_rows,
            out_path=active_path,
            sheet=cfg.active_learning_sheet,
        )


if __name__ == "__main__":
    main()
